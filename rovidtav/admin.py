# -*- coding: utf-8 -*-

import os
import codecs
from copy import copy
import StringIO
import zipfile
import datetime
import random
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from email.utils import formataddr
from _collections import defaultdict

import pytz
import pdfkit
from django import forms
from unidecode import unidecode
from django.contrib import admin
from django.contrib import messages
from django.contrib.admin.sites import AdminSite
# from django.contrib.auth import REDIRECT_FIELD_NAME
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
from django.http.response import HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from rovidtav import settings
from rovidtav.settings import WKHTMLTOPDF_EXEC
from inline_actions.admin import InlineActionsModelAdminMixin
from rovidtav.admin_helpers import ModelAdminRedirect, is_site_admin,\
    CustomDjangoObjectActions, HideIcons, SpecialOrderingChangeList,\
    DeviceOwnerListFilter, get_unread_messages_count,\
    get_unread_messages, send_assign_mail, ContentTypes, create_warehouses,\
    find_pattern, find_device_type
from rovidtav.admin_inlines import AttachmentInline, DeviceInline, NoteInline,\
    TicketInline, HistoryInline, MaterialInline, WorkItemInline,\
    TicketDeviceInline, SystemEmailInline, NTAttachmentInline, MMDeviceInline,\
    NetworkMaterialInline, NetworkWorkItemInline, PayoffTicketInline,\
    MMMaterialInline, MMAttachmentInline, WarehouseMaterialInline,\
    WarehouseDeviceInline, MMMaterialReadonlyInline, WarehouseLocationInline,\
    UninstAttachmentInline, UninstallTicketInline,\
    NTNEAttachmentInline, NTNEMaterialInline, NTNEWorkItemInline, NTNEInline,\
    IWIAttachmentInline
from rovidtav.models import Attachment, City, Client, Device, DeviceType,\
    Ticket, Note, TicketType, MaterialCategory, Material, TicketMaterial,\
    WorkItem, TicketWorkItem, Payoff, NetworkTicket, NTAttachment,\
    SystemEmail, ApplicantAttributes, DeviceOwner, Tag, Const,\
    NetworkTicketMaterial, NetworkTicketWorkItem, MaterialMovement,\
    MaterialMovementMaterial, Warehouse, WarehouseMaterial, MMAttachment,\
    DeviceReassignEvent, WarehouseLocation, UninstallTicket, UninstAttachment,\
    UninstallTicketRule, IndividualWorkItem, NetworkTicketNetworkElement,\
    NTNEType, NTNEMaterial, NTNEAttachment, NTNEWorkItem, IWIAttachment,\
    MaterialWorkitemRule
from rovidtav.forms import AttachmentForm, NoteForm, TicketMaterialForm,\
    TicketWorkItemForm, DeviceOwnerForm, TicketForm, TicketTypeForm,\
    NetworkTicketWorkItemForm, NetworkTicketMaterialForm, PayoffForm,\
    WorkItemForm, MaterialForm, MMAttachmentForm, MMMaterialForm,\
    DeviceReassignEventForm, WarehouseLocationForm, MaterialMovementForm,\
    WarehouseForm, NTNEAttachmentForm, NTNEWorkItemForm, NTNEMaterialForm,\
    NTAttachmentForm, IWIAttachmentForm, NetworkTicketForm
from rovidtav.filters import OwnerFilter, IsClosedFilter, NetworkOwnerFilter,\
    PayoffFilter, ActiveUserFilter, UninstallOwnerFilter,\
    UninstallIsClosedFilter

# ============================================================================
# MODELADMIN CLASSSES
# ============================================================================


class HideOnAdmin(admin.ModelAdmin):

    def get_model_perms(self, request):
        # Hide from admin index
        return {}


class HandleMWIOwner(object):

    """
    Handles the material and workitem owners
    """

    def _owner(self, request):
        if is_site_admin(request.user):
            return ''
        return '&owner={}'.format(request.user.pk)


class CustomUserAdmin(UserAdmin):

    list_display = ('username', 'email', 'phone_number',
                    'first_name', 'last_name',)

    def phone_number(self, obj):
        attrs = ApplicantAttributes.objects.get(user=obj)
        if attrs and attrs.tel_num:
            return attrs.tel_num
        else:
            return None

    phone_number.short_description = u'Telefonszám'


class AttachmentAdmin(HideOnAdmin, ModelAdminRedirect):

    form = AttachmentForm


class NTAttachmentAdmin(HideOnAdmin, ModelAdminRedirect):

    form = NTAttachmentForm

    def _clean(self, cell):
        """
        Splits the data by the semicolons, removes everything that is empty
        or trash data. Returns a list of useful values if more than 1,
        otherwise a single string
        """
        values = [v for v in unicode(cell.value).split(';') if v]
        return values if len(values) > 1 else values[0]

    def _get_address(self, raw_address, ticket):
        """
        Returns City object and address
        """
        if type(raw_address) == list:
            if raw_address[1].isdigit():
                offset = 1
            else:
                offset = 0
            street = raw_address[1+offset]
            house_num = u'/'.join(raw_address[2+offset:])
            address = u' '.join((street, house_num))
        else:
            address = raw_address
        return ticket.city, address

    def save_model(self, request, obj, form, change):
        if form.data.get('deviceupload'):
            xls_file = load_workbook(form.files[u'_data'].file)
            ws = xls_file.worksheets[0]
            for row in ws.iter_rows():
                try:
                    raw_address, ext_id, type_str, dev_type, _ = map(self._clean, row)
                except ValueError:
                    raw_address, ext_id, type_str, dev_type = map(self._clean, row)[:4]
                if dev_type not in ('EEP',):
                    dev_type = dev_type[0]
                city, address = self._get_address(raw_address, obj.ticket)
                dev_type_obj, _ = NTNEType.objects.get_or_create(type_str=type_str,
                                                                 type=dev_type)
                NetworkTicketNetworkElement.objects.get_or_create(
                    address=address, city=city, ticket=form.instance.ticket,
                    ext_id=ext_id, type=dev_type_obj)
            return

        super(NTAttachmentAdmin, self).save_model(request, obj, form, change)


class MMAttachmentAdmin(AttachmentAdmin):

    form = MMAttachmentForm


class NTNEAttachmentAdmin(AttachmentAdmin):

    form = NTNEAttachmentForm


class IWIAttachmentAdmin(AttachmentAdmin):

    form = IWIAttachmentForm


class MMMaterialAdmin(ModelAdminRedirect, HideOnAdmin):

    form = MMMaterialForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class PayoffAdmin(admin.ModelAdmin):

    list_display = ('full_name', 'remark')
    inlines = (PayoffTicketInline,)
    form = PayoffForm
    ordering = ('-year', '-month', '-name')

    def get_inline_instances(self, request, obj=None):
        if not obj:
            return []
        return super(PayoffAdmin, self).get_inline_instances(request, obj=None)

    def full_name(self, obj):
        return unicode(obj)


class CityAdmin(HideOnAdmin, admin.ModelAdmin):

    list_display = ('name', 'zip', 'primer', 'onuk')


class GenericHideAndRedirect(HideOnAdmin, ModelAdminRedirect):

    pass


class WarehouseLocationAdmin(HideOnAdmin, ModelAdminRedirect):

    form = WarehouseLocationForm


class TagAdmin(HideOnAdmin, admin.ModelAdmin):

    list_display = ('name', 'remark')


class DeviceOwnerAdmin(CustomDjangoObjectActions, HideOnAdmin,
                       ModelAdminRedirect, HideIcons):

    hide_add = True
    form = DeviceOwnerForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')

    def get_form(self, request, obj=None, **kwargs):
        form = super(DeviceOwnerAdmin, self).get_form(request, obj, **kwargs)
        self._hide_icons(form, ('device',))
        ticket_id = request.GET.get('ticket_id')
        user = request.user
        if ticket_id:
            ticket = Ticket.objects.get(id=int(ticket_id))
            user = ticket.owner or request.user
        try:
            warehouse = Warehouse.objects.get(owner=user)
            allowed_pks = DeviceOwner.objects.filter(
                content_type=ContentTypes.warehouse, object_id=warehouse.id).values_list('device__id', flat=True)
            devices = Device.objects.filter(id__in=allowed_pks, returned_at__isnull=True)
            form.base_fields['device'].queryset = devices
        except Warehouse.DoesNotExist:
            pass
        return form


class DeviceAdmin(CustomDjangoObjectActions,
                  ModelAdminRedirect, HideIcons):

    list_display = ('sn', 'device_type', 'owner_link',
                    'uninstall_ticket_short', 'returned_at')
    search_fields = ('type__name', 'sn')
    list_filter = (DeviceOwnerListFilter,)
    change_actions = ('new_note',)
    readonly_fields = ('returned_at',)

    inlines = (HistoryInline,)

    # change_form_template = os.path.join('rovidtav', 'select2_wide.html')
    change_list_template = os.path.join('rovidtav', 'change_list_noadd.html')

    def get_model_perms(self, request):
        # Hide from admin index
        if not request.user.is_superuser:
            return {}
        return super(DeviceAdmin, self).get_model_perms(request)

    def get_queryset(self, request):
        if request.user.is_superuser:
            return ModelAdminRedirect.get_queryset(self, request)
        wh = Warehouse.objects.get(owner=request.user.id)
        pks = DeviceOwner.objects.filter(
            content_type=ContentTypes.warehouse,
            object_id=wh.id).values_list('id', flat=True)
        return Device.objects.filter(id__in=pks).prefetch_related('uninstall_ticket')

    def device_type(self, obj):
        if obj.type:
            return obj.type.name

    device_type.short_description = u'Típus'

    def owner_link(self, obj):
        if obj.owner:
            if isinstance(obj.owner.owner, Client):
                return (u'<a href="/admin/rovidtav/client/{}/change">{}</a>'
                        u''.format(obj.owner.owner.pk, unicode(obj.owner.owner)))
            elif isinstance(obj.owner.owner, Warehouse):
                return (u'<a href="/admin/rovidtav/warehouse/{}/change">{}</a>'
                        u''.format(obj.owner.owner.pk, unicode(obj.owner.owner)))
            else:
                return obj.owner.owner

    owner_link.allow_tags = True
    owner_link.short_description = u'Tulajdonos'

    def uninstall_ticket_short(self, obj):
        if obj.uninstall_ticket:
            return u'{} (jegy {})'.format(
                obj.uninstall_ticket.client, obj.uninstall_ticket.ext_id)

    uninstall_ticket_short.short_description = u'Leszerelés jegy'

    def new_note(self, request, obj):
        returnto_tab = self.inlines.index(NoteInline)
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next=/admin/rovidtav/device/{}/change/#/tab/'
                        'inline_{}/'.format(obj.get_content_type(),
                                            obj.pk, obj.pk, returnto_tab))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def get_inline_instances(self, request, obj=None):
        if not obj:
            return []
        return super(DeviceAdmin, self).get_inline_instances(request, obj=None)


class ClientAdmin(admin.ModelAdmin):

    readonly_fields = ('name', 'city', 'address', 'mt_id', 'created_by')
    list_display = ('name', 'mt_id', 'city_name', 'address', 'created_at_fmt')
    inlines = (TicketInline, UninstallTicketInline, DeviceInline)

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        return super(ClientAdmin, self).get_inline_instances(request, obj)

    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return ('created_by',)
        return self.readonly_fields

    def city_name(self, obj):
        return u'{} ({})'.format(obj.city.name, obj.city.zip)

    city_name.short_description = u'Település'

    def created_at_fmt(self, obj):
        return obj.created_at.strftime('%Y.%m.%d %H:%M')

    created_at_fmt.short_description = u'Létrehozva'

    def get_model_perms(self, request):
        if not is_site_admin(request.user):
            return {}
        else:
            return super(ClientAdmin, self).get_model_perms(request)


class SystemEmailAdmin(admin.ModelAdmin):

    list_display = ('status', 'related_ticket', 'remark', 'created_by',
                    'created_at')
    list_filter = ('status', )

    def related_ticket(self, obj):
        rel_id = obj.content_object.pk
        return (u'<a href="/admin/rovidtav/ticket/{}/change">{}</a>'
                u''.format(rel_id, obj.content_object.ext_id))

    related_ticket.allow_tags = True
    related_ticket.short_description = u'Jegy'

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class MaterialAdmin(admin.ModelAdmin):

    list_display = ('sn', 'name', 'category', 'price', 'unit', 'comes_from',
                    'tech_display')
    search_fields = ('sn', 'name', 'category__name')
    list_filter = ('category__name', )
    form = MaterialForm

    def tech_display(self, obj):
        tech_dict = dict([(str(t[0]), t[1]) for t in Const.get_tech_choices()])
        if obj.technologies:
            return u', '.join([tech_dict[t] for t in obj.technologies])
        else:
            return u''

    tech_display.short_description = u'Technológia'


class TicketMaterialAdmin(GenericHideAndRedirect):

    form = TicketMaterialForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class NetworkTicketMaterialAdmin(GenericHideAndRedirect):

    form = NetworkTicketMaterialForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class WorkItemAdmin(admin.ModelAdmin):

    form = WorkItemForm

    list_display = ('art_number', 'name', 'art_price', 'bulk_price',
                    'given_price', 'tech_display')

    def tech_display(self, obj):
        tech_dict = dict([(str(t[0]), t[1]) for t in Const.get_tech_choices()])
        if obj.technologies:
            return u', '.join([tech_dict[t] for t in obj.technologies])
        else:
            return u''

    tech_display.short_description = u'Technológia'


class DeviceTypeAdmin(CustomDjangoObjectActions,
                      InlineActionsModelAdminMixin,
                      admin.ModelAdmin):

    list_display = ('name', 'technology', 'function', 'sn_pattern')
    ordering = ('name',)
    actions = ('refresh_pattern','apply_on_devices')
    # inlines = [DeviceTypeDeviceInline]

    def refresh_pattern(self, request, queryset):
        for device_type in queryset:
            patt = find_pattern(device_type)
            if patt:
                device_type.sn_pattern = patt
                device_type.save()

    refresh_pattern.short_description = u'SN minta felismerése'

    def apply_on_devices(self, request, queryset):
        found_types = 0
        for device_type in queryset:
            if device_type.sn_pattern:
                for device in Device.objects.filter(type__isnull=True):
                    found_types += find_device_type(device, device_type)
        messages.add_message(request, messages.INFO, u'{} eszköz típus hozzárendelve'.format(found_types))

    apply_on_devices.short_description = u'Futtatás típus nélküli eszközökre'


class TicketWorkItemAdmin(GenericHideAndRedirect):

    form = TicketWorkItemForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class NetworkTicketWorkItemAdmin(GenericHideAndRedirect):

    form = NetworkTicketWorkItemForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class NTNEWorkItemAdmin(GenericHideAndRedirect):

    form = NTNEWorkItemForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class NTNEMaterialAdmin(GenericHideAndRedirect):

    form = NTNEMaterialForm
    change_form_template = os.path.join('rovidtav', 'select2_wide.html')


class NoteAdmin(GenericHideAndRedirect):

    form = NoteForm


class TicketTypeAdmin(GenericHideAndRedirect):

    form = TicketTypeForm


class DeviceReassignEventAdmin(GenericHideAndRedirect):

    form = DeviceReassignEventForm
    add_form_template = os.path.join('rovidtav', 'device_quickadd.html')

# =============================================================================
# ADMIN PAGES
# =============================================================================


class MaterialMovementAdmin(CustomDjangoObjectActions,
                            InlineActionsModelAdminMixin,
                            admin.ModelAdmin):
    list_per_page = 50
    list_display_links = None
    list_display = ('from_to', 'mm_link', 'created', 'materials_count',
                    'devices_count', 'fin_icon')
    list_filter = ('source', 'target', 'finalized')

    inlines = [MMMaterialInline, MMDeviceInline, MMAttachmentInline,
               NoteInline]
    change_actions = ['finalize', 'new_material', 'new_device',
                      'new_attachment', 'new_note', 'uninstall_report_telekom',
                      'uninstall_report_rovidtav', 'device_summary']
    add_form_template = os.path.join('rovidtav', 'select2.html')
    readonly_fields = ['delivery_num', 'created_at']
    form = MaterialMovementForm

    def has_delete_permission(self, request, obj=None):
        return False

    def get_changelist(self, request, **kwargs):
        create_warehouses()
        return admin.ModelAdmin.get_changelist(self, request, **kwargs)

    def get_actions(self, request):
        actions = super(MaterialMovementAdmin, self).get_actions(request)
        del actions['delete_selected']
        return actions

    def get_change_actions(self, request, object_id, form_url):
        if not object_id:
            return []
        mm = MaterialMovement.objects.get(id=object_id)
        if not mm.finalized:
            actions = ['finalize']
        else:
            actions = []
        if is_site_admin(request.user) and not mm.finalized:
            actions.extend([act for act in self.change_actions
                            if act not in ('finalize',)])
        else:
            actions.extend(['new_note', 'uninstall_report_telekom',
                            'uninstall_report_rovidtav', 'device_summary'])
        return actions

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        orig_inlines = copy(self.inlines)
        if obj and obj.finalized:
            self.inlines.remove(MMMaterialInline)
            self.inlines.insert(0, MMMaterialReadonlyInline)
        instances = super(MaterialMovementAdmin, self).get_inline_instances(request, obj=None)
        self.inlines = orig_inlines
        return instances

    def _exp_act_devices(self, obj, ticket, reassign_evts):

        def is_card(dev):
            return dev.type and 'SMART CARD' in dev.type.name

        def sort(devices):
            return sorted(devices, key=lambda dev: dev.sn)

        exp_devices = Device.objects.filter(
            uninstall_ticket=ticket,
            status__in=(Const.DeviceStatus.TO_UNINSTALL,
                        Const.DeviceStatus.UNINSTALLED))
        exp_cards = [dev for dev in exp_devices if is_card(dev)]
        exp_devices = [dev for dev in exp_devices if not is_card(dev)]
        act_devices = [evt.device for evt in reassign_evts.filter(
            device__uninstall_ticket=ticket)]
        act_cards = [dev for dev in act_devices if is_card(dev)]
        act_devices = [dev for dev in act_devices if not is_card(dev)]
        return sort(exp_cards), sort(exp_devices), \
            sort(act_cards), sort(act_devices)

    def _write_row(self, row_idx, worksheet, ticket,
                   exp_cards, exp_devices, act_cards, act_devices,
                   startcol, exp_remote):
        worksheet.row_dimensions[row_idx].height = 12
        worksheet.cell(column=startcol, row=row_idx,
                       value=ticket.created_at.strftime('%Y-%m-%d'))
        worksheet.cell(column=startcol+1, row=row_idx, value=ticket.client.mt_id)
        worksheet.cell(column=startcol+2, row=row_idx, value=ticket.ext_id)
        worksheet.cell(column=startcol+3, row=row_idx, value=ticket.client.name)
        worksheet.cell(column=startcol+4, row=row_idx, value=ticket.client.address)
        exp_cards_copy = copy(exp_cards)
        act_cards_copy = copy(act_cards)
        remote_devices = (Const.DeviceFunction.BOX_IPTV,
                          Const.DeviceFunction.BOX_SAT)

        exp_idx_multi = 3 if exp_remote else 2

        for idx, dev in enumerate(exp_devices):
            worksheet.cell(column=startcol+6+idx*exp_idx_multi, row=row_idx,
                           value=dev.sn)
            if exp_remote and dev.type.function in remote_devices:
                worksheet.cell(column=startcol+8+idx*exp_idx_multi, row=row_idx,
                               value='Igen')
            try:
                if dev.type and dev.type.technology == Const.SAT:
                    card = exp_cards_copy.pop()
                    worksheet.cell(column=startcol+7+idx*exp_idx_multi, row=row_idx,
                                   value=card.sn)
            except IndexError:
                continue

        date_collected_col = 29 if exp_remote else 15
        worksheet.cell(column=date_collected_col, row=row_idx,
                       value=ticket.date_collected or ticket.closed_at)
        act_devices_startcol = 30 if exp_remote else 16
        for idx, dev in enumerate(act_devices):
            worksheet.cell(column=act_devices_startcol+idx*4, row=row_idx,
                           value=dev.sn)
            worksheet.cell(column=act_devices_startcol+3+idx*4, row=row_idx,
                           value='Igen')
            if dev.type.function in remote_devices:
                worksheet.cell(column=act_devices_startcol+2+idx*4, row=row_idx,
                               value='Igen')
            try:
                if dev.type and dev.type.technology == Const.SAT:
                    card = act_cards_copy.pop()
                    worksheet.cell(column=act_devices_startcol+1+idx*4, row=row_idx,
                                   value=card.sn)
            except IndexError:
                continue

    def _write_rovidtav_row(self, row_idx, worksheet, ticket,
                            exp_cards, exp_devices, act_cards, act_devices):
        self._write_row(row_idx, worksheet, ticket, exp_cards, exp_devices,
                        act_cards, act_devices, 11, True)

    def _write_telekom_row(self, row_idx, worksheet, ticket,
                           exp_cards, exp_devices, act_cards, act_devices):
        self._write_row(row_idx, worksheet, ticket, exp_cards, exp_devices,
                        act_cards, act_devices, 1, False)

    def _uninstall_report(self, request, template, writerow_func, obj, row_idx):
        workbook = load_workbook(template)
        worksheet = workbook.get_sheet_by_name('sikeresek')
        devices_processed = set()
        reassign_evts = obj.devicereassignevent_set.all() \
            .prefetch_related('device') \
            .prefetch_related('device__uninstall_ticket') \
            .prefetch_related('device__uninstall_ticket__client')
        for dre in reassign_evts:
            if dre.device.id in devices_processed:
                continue
            if not dre.device.uninstall_ticket:
                continue
            ticket = dre.device.uninstall_ticket
            if ticket.status not in (
                    Const.TicketStatus.DONE_SUCC,
                    Const.TicketStatus.DONE_UNSUCC):
                ticket.status = Const.TicketStatus.DONE_SUCC
                ticket.save(user=request.user)
            exp_cards, exp_devices, act_cards, act_devices = \
                self._exp_act_devices(obj, ticket, reassign_evts)
            writerow_func(
                row_idx, worksheet, ticket, exp_cards,
                exp_devices, act_cards, act_devices)
            devices_processed = devices_processed.union(
                set([dev.id for dev in act_cards + act_devices +
                     exp_cards + exp_devices]))
            row_idx += 1

        # ws.row_dimensions[row_idx].height = 12
        data = save_virtual_workbook(workbook)

        response = HttpResponse(
            content=data,
            content_type='application/vnd.ms-excel', status=200)
        response['Content-Disposition'] = ('attachment; filename=leszereles_{}'
                                           '.xlsx'.format(obj.delivery_num))
        return response

    def uninstall_report_rovidtav(self, request, obj):
        template = os.path.join(settings.BASE_DIR, 'rovidtav', 'templates',
                                'leszereles_rovidtav.xlsx')
        return self._uninstall_report(request, template,
                                      self._write_rovidtav_row,
                                      obj, row_idx=6)

    uninstall_report_rovidtav.label = u'Leszerelés xls (rövidtáv)'

    def uninstall_report_telekom(self, request, obj):
        template = os.path.join(settings.BASE_DIR, 'rovidtav', 'templates',
                                'leszereles.xlsx')
        return self._uninstall_report(request, template,
                                      self._write_telekom_row,
                                      obj, row_idx=3)

    uninstall_report_telekom.label = u'Leszerelés xls (Telekom)'

    def device_summary(self, request, obj):
        workbook = Workbook()
        worksheet = workbook.get_active_sheet()
        d = defaultdict(int)
        for e in obj.devicereassignevent_set.all():
            if e.device.type:
                d[e.device.type.name] += 1
            else:
                d[u'Ismeretlen'] += 1

        worksheet.cell(column=1, row=1, value=u'Típus')
        worksheet.cell(column=2, row=1, value=u'Darabszám')
        row_idx = 2
        for dev_type, dev_count in d.items():
            tcell = worksheet.cell(column=1, row=row_idx, value=dev_type)
            worksheet.cell(column=2, row=row_idx, value=dev_count)
            worksheet.column_dimensions[tcell.column].width = 50
            row_idx += 1

        row_idx += 2
        worksheet.cell(column=2, row=row_idx, value=u'Szériaszám')
        worksheet.cell(column=1, row=row_idx, value=u'Típus')
        for e in obj.devicereassignevent_set.all():
            row_idx += 1
            worksheet.cell(column=2, row=row_idx, value=e.device.sn)
            if e.device.type:
                t = e.device.type.name
            else:
                t = u'Ismeretlen'
            worksheet.cell(column=1, row=row_idx, value=t)

        # ws.row_dimensions[row_idx].height = 12
        data = save_virtual_workbook(workbook)

        response = HttpResponse(
            content=data,
            content_type='application/vnd.ms-excel', status=200)
        response['Content-Disposition'] = ('attachment; filename=osszesito_{}'
                                           '.xlsx'.format(obj.delivery_num))
        return response

    device_summary.label = u'Eszköz összesítő'

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def new_attachment(self, request, obj):
        return redirect('/admin/rovidtav/mmattachment/add/?materialmovement'
                        '={}&next={}'.format(
                            obj.pk, self._returnto(obj, MMAttachmentInline)))

    new_attachment.label = u'File'
    new_attachment.css_class = 'addlink'

    def new_material(self, request, obj):
        return redirect('/admin/rovidtav/materialmovementmaterial/add/?'
                        'materialmovement={}&next={}'
                        ''.format(obj.pk, self._returnto(obj, MMMaterialInline)))

    new_material.label = u'Anyag'
    new_material.css_class = 'addlink'

    def new_device(self, request, obj):
        return redirect('/admin/rovidtav/devicereassignevent/add/?materialmovement={}&next={}'
                        ''.format(obj.pk, self._returnto(obj, MMDeviceInline)))

    new_device.label = u'Eszköz'
    new_device.css_class = 'addlink'

    def finalize(self, request, obj):

        def _substract(warehouse, movement):
            to_go = movement.amount
            for material in WarehouseMaterial.objects.filter(
                    material=movement.material, warehouse=warehouse):
                if movement.amount < material.amount:
                    material.amount -= movement.amount
                    material.save()
                    break
                elif material.amount < to_go:
                    to_go -= movement.amount
                    material.delete()

            if to_go:
                # Some elements remained and could not be sobstracted
                # ADD MESSAGE?
                pass

            return to_go

        for movement in MaterialMovementMaterial.objects.filter(materialmovement=obj):
            # Substract from source
            _substract(obj.source, movement)

            # Add to target
            material = WarehouseMaterial.objects.filter(
                material=movement.material, warehouse=obj.target)
            if not material:
                WarehouseMaterial.objects.create(
                    material=movement.material, warehouse=obj.target,
                    amount=movement.amount, created_by=request.user,
                    location=movement.location_to)
            else:
                material = random.choice(material)
                material.amount += movement.amount
                if movement.location_to:
                    material.location = movement.location_to
                material.save()

        to_warehouse = obj.target.owner is None
        for dre in DeviceReassignEvent.objects.filter(materialmovement=obj):
            if to_warehouse:
                dre.device.end_life()
            else:
                dre.device.start_clean()

            try:
                device_owner = DeviceOwner.objects.get(device=dre.device)
                device_owner.content_type = ContentTypes.warehouse
                device_owner.object_id = obj.target.id
                device_owner.save(user=request.user)
            except DeviceOwner.DoesNotExist:
                device_owner = DeviceOwner.objects.create(
                    device=dre.device, content_type=ContentTypes.warehouse,
                    object_id=obj.target.id)

        obj.finalized = True
        obj.save()
        messages.add_message(request, messages.INFO,
                             u'{} véglegesítve'.format(obj.delivery_num))
        return redirect('/admin/rovidtav/materialmovement')

    finalize.label = u'Véglegesít'
    finalize.onclick = u"return confirm('Biztos?')"
    finalize.allow_tags = True

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/materialmovement/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    # =========================================================================
    # FIELDS
    # =========================================================================

    def from_to(self, obj):
        return (u'{} <img style="width: 15px; height: 10px" src="/static/images'
                u'/arrow_in.png" /> {}'.format(obj.source, obj.target))

    from_to.allow_tags = True
    from_to.short_description = u'Irány'

    def created(self, obj):
        return u'{} - {}'.format(obj.created_by, obj.created_at.strftime('%Y-%m-%d'))

    created.short_description = u'Rögzítette'

    def materials_count(self, obj):
        return MaterialMovementMaterial.objects.filter(materialmovement=obj).count()

    materials_count.short_description = u'Anyagok'

    def devices_count(self, obj):
        return DeviceReassignEvent.objects.filter(materialmovement=obj).count()

    devices_count.short_description = u'Eszközök'

    def mm_link(self, obj):
        return (u'<a href="/admin/rovidtav/materialmovement/{}/change#/tab/inline_0/">'
                u'{}</a>'.format(obj.pk, obj.delivery_num))

    mm_link.allow_tags = True
    mm_link.short_description = u'Szállító száma'

    def fin_icon(self, obj):
        if obj.finalized:
            return u'<span style="color: #03A101">&#10003;</span>'
        else:
            return u'<span style="color: #A10101">&#10007;</span>'

    fin_icon.allow_tags = True
    fin_icon.short_description = u'Véglegesítve'


class WarehouseAdmin(CustomDjangoObjectActions,
                     InlineActionsModelAdminMixin,
                     admin.ModelAdmin):

    inlines = [WarehouseMaterialInline, NoteInline, WarehouseDeviceInline,
               WarehouseLocationInline]
    form = WarehouseForm
    list_display = ['warehouse_name', 'num_devices', 'num_materials']
    list_filter = [ActiveUserFilter]
    change_actions = ['new_note', 'new_location']
    ordering = ['owner__last_name', 'owner__first_name', 'name']

    def __init__(self, model, admin_site):
        admin.ModelAdmin.__init__(self, model, admin_site)
        self.deviceowner_ct = ContentTypes.deviceowner

    def get_changelist(self, request, **kwargs):
        create_warehouses()
        return admin.ModelAdmin.get_changelist(self, request, **kwargs)

    def has_delete_permission(self, request, obj=None):
        return False

    def get_actions(self, request):
        actions = super(WarehouseAdmin, self).get_actions(request)
        if 'delete_selected' in actions:
            del actions['delete_selected']
        return actions

    def _add_device_summary(self, obj):
        """
        Adds a summary of the device type counts to the general tab as
        readonly fields
        """
        dev_counts = defaultdict(int)
        for dev in DeviceOwner.objects.filter(
                content_type=ContentTypes.warehouse, object_id=obj.id):
            dev_counts[dev.device.type or u'Egyéb eszköz'] += 1
        for cnt_key, cnt in dev_counts.items():
            self.form.declared_fields[unicode(cnt_key)] = forms.CharField(
                required=False, disabled=True, initial=cnt,
                widget=forms.TextInput(attrs={'style': 'color: #A44'}))

        return sorted(map(unicode, dev_counts.keys()))

    def get_readonly_fields(self, request, obj=None):
        if obj:
            fields = ['name', 'city', 'address']
        else:
            fields = []

        fields += (list(self.readonly_fields) or [])
        return fields

    def get_fields(self, request, obj=None):
        if obj:
            extra_fields = self._add_device_summary(obj)
        else:
            extra_fields = []
        fields = ['owner', 'name', 'city', 'address']
        if obj:
            fields.remove('owner')
            fields.remove('city')
            fields.remove('address')
        return fields + extra_fields

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        return super(WarehouseAdmin, self).get_inline_instances(request, obj)

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def new_location(self, request, obj):
        return redirect('/admin/rovidtav/warehouselocation/add/?warehouse='
                        '{}&next={}'.format(
                            obj.pk,
                            self._returnto(obj, WarehouseLocationInline)))

    new_location.label = u'Raktár hely'
    new_location.css_class = 'addlink'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/warehouse/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    def warehouse_name(self, obj):
        if obj.owner:
            return obj.name
        return u'Raktár - {} ({})'.format(obj.name, obj.city.name)

    warehouse_name.short_description = u''

    def num_devices(self, obj):
        return DeviceOwner.objects.filter(
            content_type=ContentTypes.warehouse, object_id=obj.id).count()

    num_devices.short_description = u'Eszközök'

    def num_materials(self, obj):
        return WarehouseMaterial.objects.filter(warehouse=obj).count()

    num_materials.short_description = u'Anyagok'


class _TicketFields(object):

    def client_link(self, obj):
        return ('<a href="/admin/rovidtav/client/{}/change">{}</a>'
                ''.format(obj.client.pk, obj.client.mt_id))

    client_link.allow_tags = True
    client_link.short_description = u'Ügyfél'

    def client_mt_id(self, obj):
        return obj.client.mt_id

    client_mt_id.short_description = u'MT ID'

    def full_address(self, obj):
        return u'{} {}, {}'.format(obj.city.zip, obj.city.name,
                                   obj.address)

    full_address.short_description = u'Cím'

    def client_name(self, obj):
        return obj.client.name

    client_name.short_description = u'Ügyfél neve'
    client_name.admin_order_field = 'client__name'

    def client_phone(self, obj):

        def _fmt(phone_num):
            if not phone_num.startswith('+36') and phone_num:
                if phone_num.startswith('06') or \
                        phone_num.startswith('36'):
                    phone_num = phone_num[2:]
                phone_num = '+36' + phone_num
            return '<a href="tel:{}">{}</a>'.format(phone_num, phone_num)

        return '</p><label></label><p>'.join(
            map(_fmt, map(unicode.strip, obj.client.phone.split(','))))

    client_phone.short_description = u'Telefonszám'
    client_phone.allow_tags = True

    def primer(self, obj):
        return obj.city.primer

    primer.short_description = u'Primer'
    primer.admin_order_field = 'city__primer'

    def city_name(self, obj):
        return u'{} {}'.format(obj.city.name, obj.city.zip)

    city_name.short_description = u'Település'
    city_name.admin_order_field = 'city__name'

    def created_at_fmt(self, obj):
        # return obj.created_at.strftime('%Y.%m.%d %H:%M')
        return obj.created_at.strftime('%Y.%m.%d')

    created_at_fmt.short_description = u'Rögz.'
    created_at_fmt.admin_order_field = ('created_at')

    def agreed_time_fmt(self, obj):
        # return obj.created_at.strftime('%Y.%m.%d %H:%M')
        result = None
        if obj.agreed_time_from:
            result = obj.agreed_time_from.astimezone(pytz.timezone("Europe/Budapest")).strftime('%m.%d %H')
        if obj.agreed_time_to:
            result += '-' + obj.agreed_time_to.astimezone(pytz.timezone("Europe/Budapest")).strftime('%H')
        return result

    agreed_time_fmt.short_description = u'Egyzt. idő'
    agreed_time_fmt.admin_order_field = ('agreed_time_from')

    def closed_at_fmt(self, obj):
        # return obj.created_at.strftime('%Y.%m.%d %H:%M')
        return obj.closed_at.strftime('%Y.%m.%d') if obj.closed_at else None

    closed_at_fmt.short_description = u'Lezárva'
    closed_at_fmt.admin_order_field = ('closed_at')


class IndividualWorkItemAdmin(CustomDjangoObjectActions,
                              InlineActionsModelAdminMixin,
                              admin.ModelAdmin, HideIcons):

    list_display = ('owner', 'work_date_fmt', 'price', 'remarks_short')
    inlines = (IWIAttachmentInline,)
    change_actions = ['new_attachment']

    def has_delete_permission(self, request, obj=None):
        return False

    def get_form(self, request, obj=None, **kwargs):
        form = admin.ModelAdmin.get_form(self, request, obj=obj, **kwargs)
        if not is_site_admin(request.user) and not obj:
            form.base_fields['owner'].initial = request.user
            form.base_fields['owner'].widget = forms.HiddenInput()
        return form

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        return super(IndividualWorkItemAdmin, self).get_inline_instances(request, obj)

    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return []
        if not is_site_admin(request.user):
            return ('price', 'owner')
        return []

    def get_list_filter(self, request):
        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return ('owner', 'created_at')
            else:
                return []

    def get_queryset(self, request):
        qs = admin.ModelAdmin.get_queryset(self, request)
        if is_site_admin(request.user):
            return qs
        return qs.filter(owner=request.user)

    def work_date_fmt(self, obj):
        return obj.work_date.strftime('%Y.%m.%d')

    work_date_fmt.short_description = u'Munka dátuma'

    def remarks_short(self, obj):
        return obj.remark[:25].strip() + u'...' \
            if len(obj.remark) > 25 else obj.remark

    remarks_short.short_description = u'Megjegyzés'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/individualworkitem/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    def new_attachment(self, request, obj):
        return redirect('/admin/rovidtav/iwiattachment/add/?work_item'
                        '={}&next={}'.format(
                            obj.pk, self._returnto(obj, IWIAttachmentInline)))

    new_attachment.label = u'File'
    new_attachment.css_class = 'addlink'


class TicketAdmin(CustomDjangoObjectActions,
                  InlineActionsModelAdminMixin,
                  admin.ModelAdmin,
                  HideIcons, _TicketFields, HandleMWIOwner):

    # =========================================================================
    # PARAMETERS
    # =========================================================================
    form = TicketForm
    add_form_template = os.path.join('rovidtav', 'select2.html')
    # change_form_template = os.path.join('admin',
    #                                    'two_column_change_form.html')

    list_per_page = 200
    list_display_links = None
    list_display = ('ext_id_link', 'address', 'city_name', 'client_name',
                    # 'client_link',
                    'ticket_type', 'created_at_fmt',
                    'closed_at_fmt', 'owner', 'status', 'agreed_time_fmt',
                    'primer', 'has_images_nice', 'collectable', 'remark',
                    'payoff_link', 'ticket_tags_nice')
    # TODO: check if this is useful
    # list_editable = ('owner', )
    search_fields = ('client__name', 'client__mt_id', 'city__name',
                     'city__zip', 'ext_id', 'address', 'remark')
    change_actions = ('new_note', 'new_attachment', 'new_material',
                      'new_device', 'new_workitem', 'download_html')
    changelist_actions = ('summary_list',)
    inlines = (NoteInline, AttachmentInline, MaterialInline,
               WorkItemInline, TicketDeviceInline, SystemEmailInline)
    ordering = ('-created_at',)
    fields = ['ext_id', 'client', 'ticket_types', 'city', 'address',
              'client_phone', 'owner', 'status', 'closed_at',
              'agreed_time_from', 'agreed_time_to',
              'remark', 'ticket_tags', 'payoffs', 'collectable', 'created_at', ]
    readonly_fields = ('client_phone', 'full_address', 'collectable',
                       'agreed_time_from', 'agreed_time_to')
    exclude = ['additional', 'created_by']
    actions = ['download_action']

    class Media:
        css = {
            'all': ('css/ticket.css',)
        }
        js = ('js/ticket_list.js',)

    # =========================================================================
    # METHOD OVERRIDES
    # =========================================================================

    def get_form(self, request, obj=None, **kwargs):
        form = super(TicketAdmin, self).get_form(request, obj, **kwargs)
        if obj and is_site_admin(request.user):
            self._hide_icons(form, ('payoffs',), show_add=True)
            self._hide_icons(form, ('owner',))
        return form

    def get_changelist(self, request, **kwargs):
        return SpecialOrderingChangeList

    def changeform_view(self, request, object_id=None, form_url='',
                        extra_context=None):
        extra_context = extra_context or {}
        extra_context['siteSpecificContext'] = {
            'closed_statuses': [Const.TicketStatus.DONE_SUCC,
                                Const.TicketStatus.DONE_UNSUCC,
                                Const.TicketStatus.DUPLICATE]
        }
        extra_context['hideSaveOnTabs'] = 'true'

        return InlineActionsModelAdminMixin.changeform_view(self, request, object_id=object_id, form_url=form_url, extra_context=extra_context)

    def changelist_view(self, request, extra_context=None):
        """
        We need to remove the links to the client and the payoff if the user
        is not an admin
        """
        response = super(TicketAdmin, self).changelist_view(request,
                                                            extra_context)
        if not is_site_admin(request.user):
            subst = {'client_link': 'client_mt_id',
                     }
            exclude = ['payoff_link']
            columns = response.context_data['cl'].list_display
            new_cols = [subst.get(c, c) for c in columns if c not in exclude]
            response.context_data['cl'].list_display = new_cols
        return response

    def get_actions(self, request):
        actions = super(TicketAdmin, self).get_actions(request)
        del actions['delete_selected']
        return actions

    def download_action(self, request, queryset):
        temp = StringIO.StringIO()
        with zipfile.ZipFile(temp, 'w') as archive:
            for ticket in queryset:
                attachments = []
                for att in Attachment.objects.filter(ticket=ticket):
                    if (att.is_image() and not att.name.lower().startswith('imdb')) or \
                             att.name.endswith('.pdf'):
                        attachments.append(att)
                for att in attachments:
                    archive.writestr('{}/{}'.format(ticket.ext_id, unidecode(att.name)),
                                     att.data)

        temp.seek(0)
        response = HttpResponse(temp,
                                content_type='application/force-download')
        fname = datetime.datetime.now().strftime('jegyek_%y%m%d%H%M.zip')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(fname)
        return response

    download_action.short_description = u'Letöltés ZIP-ben'

    def get_change_actions(self, request, object_id, form_url):
        obj = Ticket.objects.get(pk=object_id)
        if is_site_admin(request.user) or \
                obj.status in (u'Kiadva', u'Folyamatban'):
            actions = ('new_note', 'new_attachment',
                       'new_material', 'new_workitem',
                       'new_device')
            if obj.is_install_ticket():
                actions += ('download_html',)
            return actions
        else:
            return ('new_note',)

    def get_list_filter(self, request):
        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return (
                        'city__primer', OwnerFilter, IsClosedFilter,
                        'has_images', 'ticket_tags', PayoffFilter,
                        )
            else:
                return (IsClosedFilter,)

    def lookup_allowed(self, key, value):
        if key in ('city__primer'):
            return True
        return super(TicketAdmin, self).lookup_allowed(key, value)

    def get_queryset(self, request):
        qs = super(TicketAdmin, self).get_queryset(request)
        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return qs
            return qs.filter(owner=request.user)
        return qs

    def get_readonly_fields(self, request, obj=None):
        if obj:
            fields = ('created_by', 'created_at', 'ext_id', 'client',
                      'ticket_types', 'city', 'address')
        else:
            fields = ()
        fields += (self.readonly_fields or tuple())
        if not is_site_admin(request.user):
            fields += ('owner', 'payoffs', 'remark', 'ticket_tags')
            if obj.status not in (u'Kiadva', u'Folyamatban'):
                fields += ('status', 'closed_at')
        return fields

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        orig_inlines = self.inlines
        if not is_site_admin(request.user):
            # Remove the email inline if not an admin
            self.inlines = [i for i in self.inlines
                            if i not in (SystemEmailInline, )]
        inline_inst = super(TicketAdmin, self).get_inline_instances(request,
                                                                    obj=None)
        self.inlines = orig_inlines
        return inline_inst

    def save_model(self, request, obj, form, change):
        """
        Saves the model and handles notifications if needed
        """
        notify = obj.save()
        if notify and obj.owner and obj.owner.email:
            try:
                ticket_html = Attachment.objects.get(
                    ticket=obj, name='Hibajegy.html').data
            except Attachment.DoesNotExist:
                ticket_html = ''
            ticket_url = ('{}/admin/rovidtav/ticket/{}'
                          ''.format(settings.SELF_URL, obj.pk))

            html_maxlen = 100000
            attachment = None
            if ticket_html and len(ticket_html) > html_maxlen:
                try:
                    attachment = \
                        MIMEText(ticket_html.encode('utf-8'), 'html', 'UTF-8')
                except Exception:
                    attachment = MIMEText(ticket_html, 'html', 'UTF-8')
                attachment.add_header(
                    u'Content-Disposition',
                    u'attachment; filename=hibajegy.html',
                )
            ctx = {'ticket_url': ticket_url,
                   'ticket_html': '' if attachment else ticket_html,
                   'ticket_too_long': bool(attachment)}

            msg = MIMEMultipart()
            msg['Subject'] = u'Új jegy - {} {} - Task Nr: {}'.format(
                obj.city.name, obj.address, obj.ext_id)
            msg_from = formataddr((str(Header(u'Rövidtáv rendszer', 'utf-8')), settings.EMAIL_SENDER))
            msg['From'] = msg_from
            msg['To'] = obj.owner.email
            html_template = render_to_string('assign_notification.html',
                                             context=ctx)
            msg.attach(MIMEText(html_template, 'html', 'utf-8'))

            if attachment:
                msg.attach(attachment)

            send_assign_mail(msg, obj)

    # =========================================================================
    # FIELDS
    # =========================================================================

    def ext_id_link(self, obj):
        return (u'<a href="/admin/rovidtav/ticket/{}/change#/tab/inline_0/">'
                u'{}</a>'.format(obj.pk, obj.ext_id))

    ext_id_link.allow_tags = True
    ext_id_link.short_description = u'Jegy ID'

    def collectable(self, obj):
        return obj[Ticket.Keys.COLLECTABLE_MONEY] or '-'

    collectable.short_description = u'Beszed.'

    def payoff_link(self, obj):
        payoffs = []
        if obj.payoffs:
            for payoff in obj.payoffs.all():
                payoff_a = (u'<a href="/admin/rovidtav/payoff/{}/change">{}'
                            u'</a>'.format(payoff.pk, unicode(payoff)))
                payoffs.append(payoff_a)
            return u',&nbsp'.join(payoffs)
        else:
            return None

    payoff_link.allow_tags = True
    payoff_link.short_description = u'Elszám.'

    def ticket_type(self, obj):
        types = ' / '.join([t.name for t in obj.ticket_types.all()])
        return types[:25].strip() + u'...' if len(types) > 25 else types

    ticket_type.short_description = u'Tipus'
    # ticket_type.admin_order_field = ('created_at')

    def has_images_nice(self, obj):
        return u'✓' if obj.has_images else ''

    has_images_nice.short_description = u'Kép'

    def ticket_tags_nice(self, obj):
        return ', '.join([t.name for t in obj.ticket_tags.all()])

    ticket_tags_nice.short_description = u'Cimkék'

    # =========================================================================
    # ACTIONS
    # =========================================================================

    def summary_list(self, request, obj):
        return redirect('/reports/osszesito')

    summary_list.label = u'Összesítő lista'
    # summary_list.css_class = 'addlink'

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def new_attachment(self, request, obj):
        return redirect('/admin/rovidtav/attachment/add/?ticket={}&next={}'
                        ''.format(obj.pk,
                                  self._returnto(obj, AttachmentInline)))

    new_attachment.label = u'File'
    new_attachment.css_class = 'addlink'

    def new_material(self, request, obj):
        return redirect('/admin/rovidtav/ticketmaterial/add/?ticket={}{}&next={}'
                        ''.format(obj.pk, self._owner(request),
                                  self._returnto(obj, MaterialInline)))

    new_material.label = u'Anyag'
    new_material.css_class = 'addlink'

    def new_workitem(self, request, obj):
        return redirect('/admin/rovidtav/ticketworkitem/add/?ticket={}{}&next={}'
                        ''.format(obj.pk, self._owner(request),
                                  self._returnto(obj, WorkItemInline)))

    new_workitem.label = u'Munka'
    new_workitem.css_class = 'addlink'

    def new_device(self, request, obj):
        return redirect('/admin/rovidtav/deviceowner/add/?content_type={}'
                        '&object_id={}&ticket_id={}&next={}'
                        ''.format(obj.client.get_content_type(), obj.client.pk,
                                  obj.pk,
                                  self._returnto(obj, TicketDeviceInline)))

    new_device.label = u'Eszköz'
    new_device.css_class = 'addlink'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/ticket/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    def download_html(self, request, ticket):
        try:
            map_img = Attachment.objects.get(ticket=ticket,
                                             name__istartswith='imdb')
            map_img = map_img._data
        except Attachment.DoesNotExist:
            map_img = None

        materials = {}
        for material_sn in ('40292016', '40296137', '40299501', '40306778',
                            '40296139', '40292261'):
            try:
                material = Material.objects.get(sn=material_sn)
            except Material.DoesNotExist:
                continue

            try:
                tm = TicketMaterial.objects.get(ticket=ticket,
                                                material=material)
                materials[material_sn] = tm.amount
            except TicketMaterial.DoesNotExist:
                continue

        ctx = {'wfms_id': ticket.ext_id,
               'client_name': ticket.client.name,
               'city_primer': ticket.city.primer or '',
               'city_name': ticket.city.name,
               'city_zip': ticket.city.zip,
               'owner_name': u'{} {}'.format(ticket.owner.last_name,
                                             ticket.owner.first_name)
               if ticket.owner else '',
               'client_address': ticket.address,
               'map': map_img,
               'today': datetime.datetime.now().strftime('%Y-%m-%d'),
               'year': datetime.datetime.now().strftime('%Y'),
               'materials': materials,
               }
        response = render(request, 'leltaruj1.html', ctx)
        response.content_type = 'text/html'
        response['Content-Disposition'] = ('attachment; filename=leltar_'
                                           '{}.html'.format(ticket.ext_id))
        return response

    download_html.label = u'Leltár adatlap'
    download_html.css_class = 'downloadlink'


class UninstallTicketAdmin(
        CustomDjangoObjectActions, InlineActionsModelAdminMixin,
        admin.ModelAdmin, HideIcons, _TicketFields):

    # =========================================================================
    # PARAMETERS
    # =========================================================================
    add_form_template = os.path.join('rovidtav', 'select2.html')

    list_per_page = 200
    list_display_links = None
    list_display = ('ext_id_link', 'address', 'city_name', 'client_name',
                    # 'client_link',
                    'ticket_type_short', 'created_at_fmt', 'closed_at_fmt',
                    'owner', 'status', 'primer')
    # TODO: check if this is useful
    # list_editable = ('owner', )
    search_fields = ('client__name', 'client__mt_id', 'city__name',
                     'city__zip', 'ext_id', 'address', 'client__phone')
    change_actions = ('new_note', 'assign_to_me', 'uninstall_document')
    inlines = (NoteInline, UninstAttachmentInline, TicketDeviceInline,
               SystemEmailInline)
    ordering = ('-created_at',)
    fields = ['ext_id', 'client', 'ticket_type', 'city', 'address',
              'client_phone', 'owner', 'date_collected', 'status', 'closed_at',
              'ticket_tags', 'created_at', ]
    readonly_fields = ('client_phone', 'full_address', 'ticket_type',
                       'address', 'city', 'ext_id', 'client', 'created_at',
                       'closed_at')
    exclude = ['additional', 'created_by']

    #def has_delete_permission(self, request, obj=None):
    #    return False

    def lookup_allowed(self, key, value):
        if key in ('city__primer'):
            return True
        return super(UninstallTicketAdmin, self).lookup_allowed(key, value)

    def get_fields(self, request, obj=None):
        if obj:
            return self.fields
        else:
            return [
                'ext_id', 'client', 'ticket_type', 'city', 'address',
                'owner', 'date_collected', 'status', 'created_at', 'closed_at',
                'ticket_tags']

    def get_list_filter(self, request):
        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return (UninstallOwnerFilter, IsClosedFilter,
                        'ticket_tags', 'city', 'city__primer')
            else:
                return (UninstallIsClosedFilter, 'city__primer')

    def get_actions(self, request):
        if not request.user.is_superuser:
            return []

        actions = super(UninstallTicketAdmin, self).get_actions(request)

        def _assign(technician, instance, request, queryset):
            queryset.update(owner=technician)

        for technician in User.objects.filter(groups__name=u'Leszerelő'):
            action_label = u'Szerelőnek kiad: {}'.format(technician.username)
            action_name = 'assign_to_{}'.format(technician.id)
            _assign.short_description = action_label
            setattr(self, action_name,
                    lambda instance, request, queryset:
                        _assign(technician, instance, request, queryset))
            actions[action_name] = (getattr(self, action_name),
                                    action_name, action_label)
        return actions

    def get_change_actions(self, request, object_id, form_url):
        if not object_id:
            return []
        actions = ['uninstall_document', 'new_note']
        obj = UninstallTicket.objects.get(id=object_id)
        if not request.user.is_superuser and obj.owner != request.user:
            actions.insert(0, 'assign_to_me')
        return actions

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.strip('/').endswith('change'):
            return []
        orig_inlines = self.inlines
        if not is_site_admin(request.user):
            # Remove the email inline if not an admin
            self.inlines = [i for i in self.inlines
                            if i not in (SystemEmailInline, )]
        inline_inst = super(UninstallTicketAdmin, self) \
            .get_inline_instances(request, obj=None)
        self.inlines = orig_inlines
        return inline_inst

    def get_readonly_fields(self, request, obj=None):
        if obj:
            fields = self.readonly_fields
        else:
            fields = ()
        if not is_site_admin(request.user):
            fields += ('ticket_tags', 'owner')
            if obj and obj.status not in (u'Új', u'Kiadva', u'Folyamatban'):
                fields += ('status', 'closed_at',)
        return fields

    def ext_id_link(self, obj):
        return (u'<a href="/admin/rovidtav/uninstallticket/{}/change#/tab/inline_0/">'
                u'{}</a>'.format(obj.pk, obj.ext_id))

    ext_id_link.allow_tags = True
    ext_id_link.short_description = u'Jegy ID'

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def uninstall_document(self, request, obj):

        def _postfix(replacements, key_prefix):
            return str(len([k for k in replacements.keys()
                            if k.startswith(key_prefix)
                            and replacements[k]]) + 1)

        with codecs.open(os.path.join(settings.PROJECT_DIR, 'templates',
                               'uninstall_template.html')) as template:
            template_content = template.read().decode('utf-8')
        replacements = {
            'date': datetime.datetime.now().strftime('%Y.%m.%d %H:%M'),
            'wfms_id': obj.ext_id,
            'mt_id': obj.client.mt_id,
            'client_name': obj.client.name,
            'client_address': u'{} {}'.format(obj.client.city, obj.client.address),
            'client_phone': obj.client.phone,
            'year': datetime.datetime.now().strftime('%Y'),
            'card1': '',
            'card2': '',
            'card3': '',
            'stb1': '',
            'stb2': '',
            'stb3': '',
            'owned1': '',
            'owned2': '',
            'owned3': '',
            'rented_stb': 0,
        }
        for dev in [do.device for do in DeviceOwner.objects.filter(
                    content_type=ContentTypes.client,
                    object_id=obj.client.id)]:
            if 'SMART CARD' in dev.type.name:
                key = 'card' + _postfix(replacements, 'card')
                if dev.status == Const.DeviceStatus.TO_UNINSTALL:
                    replacements[key] = dev.sn
            else:
                postfix = _postfix(replacements, 'stb')
                if dev.status == Const.DeviceStatus.TO_UNINSTALL:
                    replacements['stb' + postfix] = dev.sn
                    replacements['owned' + postfix] = u'Bérelt'
                    replacements['rented_stb'] += 1

        for key, data in replacements.items():
            template_content = template_content.replace(
                u'{{' + key + u'}}', unicode(data))
        config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_EXEC)
        pdfdoc = pdfkit.from_string(template_content, False, configuration=config)
        return HttpResponse(pdfdoc,
                            content_type='application/pdf')

    uninstall_document.label = u'Leszerelési lap'

    def assign_to_me(self, request, obj):
        obj.owner = request.user
        obj.save(user=request.user)

    assign_to_me.label = u'Felveszem'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/uninstallticket/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    def ticket_type_short(self, obj):
        ttype = unicode(obj.ticket_type)
        if ttype.startswith('L-Kiemelt'):
            ttype = u'Begyűjtés ' + ttype[ttype.index('('):]
        if len(ttype) > 46:
            return ttype[:43] + u'...'
        return ttype

    ticket_type_short.short_description = u'Jegy típus'


class NetworkTicketAdmin(CustomDjangoObjectActions,
                         InlineActionsModelAdminMixin,
                         admin.ModelAdmin, HideIcons, HandleMWIOwner):

    list_per_page = 200
    list_display_links = None
    list_display = ('address_link', 'city_name', 'onu',
                    'ticket_type', 'created_at_fmt',
                    'closed_at_fmt', 'owner_display', 'status',
                    'ticket_tags_nice')
    change_actions = ('new_note', 'new_attachment', 'new_material',
                      'new_workitem', 'upload_devices')
    inlines = (NoteInline, NTAttachmentInline, NetworkMaterialInline,
               NetworkWorkItemInline, NTNEInline)
    ordering = ('-created_at',)
    search_fields = ('city__name', 'city__zip', 'address',
                     'master_sn')

    fields = ['city', 'address', 'onu', 'master_sn',
              'psu_placement', 'ticket_types',
              'ticket_tags', 'technologies', 'owner', 'status', 'closed_at']
    readonly_fields = ('full_address',)
    list_filter = ('onu', NetworkOwnerFilter, IsClosedFilter, 'ticket_types')
    actions = ['download_action']
    form = NetworkTicketForm

    class Media:
        js = ('js/network_ticket.js',)

    def get_list_filter(self, request):
        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return self.list_filter
            else:
                return ('onu', IsClosedFilter,)

    def get_readonly_fields(self, request, obj=None):
        fields = self.readonly_fields
        if obj:
            fields += ('city',)
        if not is_site_admin(request.user):
            fields += ('owner', 'address', 'ticket_types', 'ticket_tags',
                       'city', 'closed_at', 'onu', 'master_sn',
                       'psu_placement',)
            if obj.status not in (Const.TicketStatus.NEW,
                                  Const.TicketStatus.IN_PROGRESS,
                                  Const.TicketStatus.ASSIGNED):
                fields += ('status',)
        return fields

    def get_actions(self, request):
        actions = super(NetworkTicketAdmin, self).get_actions(request)
        del actions['delete_selected']
        return actions

    def get_change_actions(self, request, object_id, form_url):
        actions = CustomDjangoObjectActions.get_change_actions(
            self, request, object_id, form_url)
        if not is_site_admin(request.user):
            actions = [a for a in actions if a not in ('upload_devices',)]
        return actions

    def get_form(self, request, obj=None, **kwargs):
        form = super(NetworkTicketAdmin, self).get_form(request, obj, **kwargs)
        if is_site_admin(request.user):
            self._hide_icons(form, ('owner',))
            # self._hide_icons(form, ('city',))
        return form

    def get_queryset(self, request):
        qs = super(NetworkTicketAdmin, self).get_queryset(request) \
            .prefetch_related('networkticketnetworkelement_set') \
            .prefetch_related('networkticketnetworkelement_set__type')

        if hasattr(request, 'user'):
            if is_site_admin(request.user):
                return qs
            return qs.filter(owner__in=[request.user.pk])
        return qs

    def has_delete_permission(self, request, obj=None):
        return False

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.split('#')[0].strip('/').endswith('change'):
            return []
        return super(NetworkTicketAdmin, self).get_inline_instances(request,
                                                                    obj)

    def address_link(self, obj):
        return (u'<a href="/admin/rovidtav/networkticket/{}/change#'
                u'/tab/inline_0/">{}</a>'.format(obj.pk, obj.address))

    address_link.allow_tags = True
    address_link.short_description = u'Cím'

    def ticket_type(self, obj):
        types = ' / '.join([t.name for t in obj.ticket_types.all()])
        return types[:25].strip() + u'...' if len(types) > 25 else types

    ticket_type.short_description = u'Tipus'

    def created_at_fmt(self, obj):
        # return obj.created_at.strftime('%Y.%m.%d %H:%M')
        return obj.created_at.strftime('%Y.%m.%d')

    created_at_fmt.short_description = u'Felvéve'
    created_at_fmt.admin_order_field = ('created_at')

    def closed_at_fmt(self, obj):
        # return obj.created_at.strftime('%Y.%m.%d %H:%M')
        return obj.closed_at.strftime('%Y.%m.%d') if obj.closed_at else None

    closed_at_fmt.short_description = u'Lezárva'
    closed_at_fmt.admin_order_field = ('closed_at')

    def ticket_tags_nice(self, obj):
        return ', '.join([t.name for t in obj.ticket_tags.all()])

    ticket_tags_nice.short_description = u'Cimkék'

    def city_name(self, obj):
        return u'{} {}'.format(obj.city.name, obj.city.zip)

    city_name.short_description = u'Település'
    city_name.admin_order_field = 'city__name'

    def full_address(self, obj):
        return u'{} {}, {}'.format(obj.city.zip, obj.city.name,
                                   obj.address)

    full_address.short_description = u'Cím'

    def owner_display(self, obj):
        return u', '.join([u.username for u in obj.owner.all()])

    owner_display.short_description = u'Szerelő'

    # =========================================================================
    # ACTIONS
    # =========================================================================

    def download_action(self, request, queryset):
        temp = StringIO.StringIO()
        att_count = 0
        with zipfile.ZipFile(temp, 'w') as archive:
            for ticket in queryset:
                attachments = []
                for att in NTAttachment.objects.filter(ticket=ticket):
                    if att.is_image():
                        name = ticket.address.replace(u'/', u'-') + u'_{}'.format(att.name)
                        attachments.append((name, att))
                        att_count += 1
                elements = NetworkTicketNetworkElement.objects\
                    .filter(ticket=ticket)
                for idx, att in enumerate(
                        NTNEAttachment.objects
                        .filter(network_element__in=elements)
                        .prefetch_related('network_element')):
                    if att.is_image() or att.name.endswith(('.xls', '.xlsx')):
                        ext = att.name.split('.')[-1]
                        name = u'{}_{}_{}.{}'.format(
                            att.network_element.ext_id,
                            att.network_element.address,
                            str(idx+1), ext)
                        attachments.append((name, att))
                        att_count += 1
                for name, att in attachments:
                    archive.writestr(name, att.data)

        if att_count == 0:
            messages.add_message(request, messages.INFO, u'A kiválasztott jegyek nem tartalmaznak képeket')
            return
        fname = datetime.datetime.now().strftime('halozat_jegyek_%y%m%d%H%M.zip')
        temp.seek(0)
        response = HttpResponse(temp,
                                content_type='application/force-download')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(fname)
        return response

    download_action.short_description = u'Letöltés ZIP-ben'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/networkticket/{}/change/#/tab/inline_{}/'
                ''.format(obj.pk, returnto_tab))

    def upload_devices(self, request, obj):
        return redirect('/admin/rovidtav/ntattachment/add/?ticket={}&'
                        'deviceupload=1&next={}'
                        ''.format(obj.pk, self._returnto(obj, NTNEInline)))

    upload_devices.label = u'Elemek feltöltése'

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def new_attachment(self, request, obj):
        return redirect('/admin/rovidtav/ntattachment/add/?ticket={}&next={}'
                        ''.format(obj.pk,
                                  self._returnto(obj, NTAttachmentInline)))

    new_attachment.label = u'File'
    new_attachment.css_class = 'addlink'

    def new_material(self, request, obj):
        return redirect('/admin/rovidtav/networkticketmaterial/add/?ticket={}{}&next={}'
                        ''.format(obj.pk, self._owner(request),
                                  self._returnto(obj, NetworkMaterialInline)))

    new_material.label = u'Anyag'
    new_material.css_class = 'addlink'

    def new_workitem(self, request, obj):
        return redirect('/admin/rovidtav/networkticketworkitem/add/?ticket={}{}&next={}'
                        ''.format(obj.pk,  self._owner(request),
                                  self._returnto(obj, NetworkWorkItemInline)))

    new_workitem.label = u'Munka'
    new_workitem.css_class = 'addlink'


class NetworkTicketNetworkElementAdmin(CustomDjangoObjectActions,
                                       InlineActionsModelAdminMixin,
                                       admin.ModelAdmin, HandleMWIOwner):

    search_fields = ('ext_id', 'address', 'city__name')
    list_filter = ('city__name', 'ticket__onu', 'status')
    list_display = ('ext_id', 'full_address', 'type', 'onu', 'status')
    readonly_fields = ('created_at', 'created_by', 'ext_id', 'city', 'ticket')
    fields = ('ext_id', 'full_address', 'ticket',  'status', 'type',
              'ticket_tags')
    inlines = (NoteInline, NTNEAttachmentInline, NTNEMaterialInline,
               NTNEWorkItemInline)
    change_actions = ('new_note', 'new_attachment', 'new_material',
                      'new_workitem',)

    def get_fields(self, request, obj=None):
        if obj:
            return admin.ModelAdmin.get_fields(self, request, obj=obj)
        else:
            return [f for f in admin.ModelAdmin.get_fields(self, request, obj=obj)
                    if f not in ('full_address',)] + ['city', 'address']

    def get_queryset(self, request):
        return admin.ModelAdmin.get_queryset(self, request) \
            .prefetch_related('ticket') \
            .prefetch_related('city')

    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return ('created_at', 'created_by')
        return admin.ModelAdmin.get_readonly_fields(self, request, obj=obj) + \
            ('full_address',)

    def get_change_actions(self, request, object_id, form_url):
        if not object_id:
            return []
        return super(NetworkTicketNetworkElementAdmin, self) \
            .get_change_actions(request, object_id, form_url)

    def get_inline_instances(self, request, obj=None):
        if not obj and not request.path.split('#')[0].strip('/').endswith('change'):
            return []
        return super(NetworkTicketNetworkElementAdmin, self) \
            .get_inline_instances(request, obj)

    def full_address(self, obj):
        return u'{} {}, {}'.format(obj.city.zip, obj.city.name,
                                   obj.address)

    full_address.short_description = u'Cím'

    def onu(self, obj):
        return obj.ticket.onu

    onu.short_description = u'Onu'

    def _returnto(self, obj, inline):
        returnto_tab = self.inlines.index(inline)
        return ('/admin/rovidtav/networkticketnetworkelement/{}/change/#/tab/'
                'inline_{}/'.format(obj.pk, returnto_tab))

    def new_note(self, request, obj):
        return redirect('/admin/rovidtav/note/add/?content_type={}&object_id='
                        '{}&next={}'.format(obj.get_content_type(),
                                            obj.pk,
                                            self._returnto(obj, NoteInline)))

    new_note.label = u'Megjegyzés'
    new_note.css_class = 'addlink'

    def new_attachment(self, request, obj):
        return redirect('/admin/rovidtav/ntneattachment/add/?network_element='
                        '{}&next={}'.format(
                            obj.pk, self._returnto(obj, NTNEAttachmentInline)))

    new_attachment.label = u'File'
    new_attachment.css_class = 'addlink'

    def new_material(self, request, obj):
        return redirect('/admin/rovidtav/ntnematerial/add/?network_element={}{}&next={}'
                        ''.format(obj.pk, self._owner(request),
                                  self._returnto(obj, NTNEMaterialInline)))

    new_material.label = u'Anyag'
    new_material.css_class = 'addlink'

    def new_workitem(self, request, obj):
        return redirect('/admin/rovidtav/ntneworkitem/add/?network_element={}{}&next={}'
                        ''.format(obj.pk, self._owner(request),
                                  self._returnto(obj, NTNEWorkItemInline)))

    new_workitem.label = u'Munka'
    new_workitem.css_class = 'addlink'


class CustomAdminSite(AdminSite):

    def login(self, request, extra_context=None):
        extra_context = extra_context or {}
        return super(CustomAdminSite, self).login(request, extra_context)

    def each_context(self, request):
        ctx = super(CustomAdminSite, self).each_context(request)
        ctx.update(get_unread_messages_count(request.user))
        ctx.update(get_unread_messages(request.user))
        ctx.update({'c': 'sadf'})
        return ctx


admin.site = CustomAdminSite()

admin.site.register(User, CustomUserAdmin)
admin.site.register(Group)
admin.site.register(SystemEmail, SystemEmailAdmin)

admin.site.register(City, CityAdmin)
admin.site.register(Payoff, PayoffAdmin)
admin.site.register(Client, ClientAdmin)
admin.site.register(TicketType, TicketTypeAdmin)
admin.site.register(Tag, TagAdmin)
admin.site.register(Ticket, TicketAdmin)
admin.site.register(NetworkTicket, NetworkTicketAdmin)
admin.site.register(UninstallTicket, UninstallTicketAdmin)
admin.site.register(IndividualWorkItem, IndividualWorkItemAdmin)
admin.site.register(IWIAttachment, IWIAttachmentAdmin)
admin.site.register(UninstallTicketRule)
admin.site.register(ApplicantAttributes)
admin.site.register(MaterialWorkitemRule)
admin.site.register(Note, NoteAdmin)
admin.site.register(Attachment, AttachmentAdmin)
admin.site.register(NTAttachment, NTAttachmentAdmin)
admin.site.register(UninstAttachment, AttachmentAdmin)
admin.site.register(NetworkTicketMaterial, NetworkTicketMaterialAdmin)
admin.site.register(NetworkTicketWorkItem, NetworkTicketWorkItemAdmin)
admin.site.register(NetworkTicketNetworkElement,
                    NetworkTicketNetworkElementAdmin)
admin.site.register(NTNEType, GenericHideAndRedirect)
admin.site.register(NTNEMaterial, NTNEMaterialAdmin)
admin.site.register(NTNEWorkItem, NTNEWorkItemAdmin)
admin.site.register(NTNEAttachment, NTNEAttachmentAdmin)

admin.site.register(MaterialMovement, MaterialMovementAdmin)
admin.site.register(MaterialMovementMaterial, MMMaterialAdmin)
admin.site.register(MMAttachment, MMAttachmentAdmin)
admin.site.register(Warehouse, WarehouseAdmin)
admin.site.register(WarehouseLocation, WarehouseLocationAdmin)
admin.site.register(WarehouseMaterial, GenericHideAndRedirect)
admin.site.register(DeviceReassignEvent, DeviceReassignEventAdmin)

admin.site.register(WorkItem, WorkItemAdmin)
admin.site.register(TicketWorkItem, TicketWorkItemAdmin)

admin.site.register(Device, DeviceAdmin)
admin.site.register(DeviceOwner, DeviceOwnerAdmin)
admin.site.register(DeviceType, DeviceTypeAdmin)

admin.site.register(Material, MaterialAdmin)
admin.site.register(MaterialCategory, GenericHideAndRedirect)
admin.site.register(TicketMaterial, TicketMaterialAdmin)
